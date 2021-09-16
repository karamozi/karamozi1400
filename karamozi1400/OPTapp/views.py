from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from django.contrib.auth import authenticate, login, logout as django_logout
from django.http import HttpResponseRedirect, HttpResponse,JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from openpyxl import Workbook, load_workbook, styles
from OPTapp.models import Schools, Classes, Table
from django.core.exceptions import ValidationError
from django.contrib.auth.models import User
from OPTapp.forms import UploadFileForm
from django.core import validators
from django.conf import settings
from django.urls import reverse
import os


from itertools import product
from sys import stdout as out
from mip import *

# Create your views here.

filename = os.path.join(settings.BASE_DIR, 'static', 'Three_Schools.xlsx')
wb = load_workbook(filename)
ws = wb['Data']

schools = list()
sunday = ws['S2'].value

for i in range(1, 156):
    if ws['r' + str(i)].value:
        schools.append(ws['r' + str(i)].value)


hour_list = range(8, 19)
minute_list = ['00', '15', '30', '45']


time_list = list()
for hour in hour_list:
    for minute in minute_list:
        time = str(hour) + ':' + str(minute)
        time_list.append(time)

time_list = ['7:30','7:45'] + time_list + ['19:00']


def Home(request):
    if request.user.is_authenticated:
        return redirect('school_selection')
    else:
        return redirect('login')


def userCreate(request):

    email_message = str()
    username_message = str()
    password_message = str()
    confirm_password_message = str()
    username = str()
    email = str()
    password = str()
    confirm_password = str()

    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')

        emails = [i.email for i in User.objects.all()]
        usernames = [i.username for i in User.objects.all()]

        if not email:
            email_message = 'وارد کردن ایمیل الزامی است'
        elif email in emails:
            email_message = 'این ایمیل قبلا استفاده شده است'
        else:
            try:
                validators.validate_email(email)
            except ValidationError:
                email_message = 'ایمیل اشتباه است'

        if not username:
            username_message = 'نام کاربری الزامی است'
        elif username in usernames:
            username_message = 'این نام قبلا استفاده شده است'

        if not password:
            password_message = 'رمز عبور را وارد کنید'
        elif password != confirm_password:
            confirm_password_message = 'رمز عبور یکسان نیست'

        if email_message == '' and username_message == '' and password_message == '':
            user = User.objects.create_user(username, email, password)
            user.save()

            return redirect(reverse('login'))

    return render(request, 'signup.html', {'email_message': email_message,
                                           'username_message': username_message,
                                           'password_message': password_message,
                                           'confirm_password_message': confirm_password_message,
                                           'username': username,
                                           'email': email,
                                           'password': password,
                                           'confirm_password': confirm_password})


def login_page(request):
    message = str()
    username = str()
    password = str()
    if request.method == 'POST':
        next = request.POST.get('next')
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        if user:
            login(request, user)
            if next:
                return redirect(next)
            else:
                return redirect('school_selection')

        else:
            message = 'نام كاربري يا رمز عبور درست نمي‌باشد!!'

    return render(request, 'login.html', {'message': message, 'username': username, 'password': password})


def logout(request):
    django_logout(request)
    return redirect('login')


@login_required
def school_selection(request):
    if request.method == 'POST':
        valueList = request.POST.get('valueList')
        item = Schools.objects.filter(user=request.user).all().delete()
        if valueList:
            valueList = valueList.split(',')

            for i in valueList:
                new_item = Schools(school=i, user=request.user)
                new_item.save()
            return redirect('class_selection')
        else:
            return redirect('class_selection')
    selected_schools = [
        i.school for i in Schools.objects.filter(user=request.user)]
    remaining_schools = [i for i in schools if i not in selected_schools]

    return render(request, 'school_selection.html', {'remaining_schools': remaining_schools,
                                                     'selected_schools': selected_schools})


@login_required
def class_selection(request):
    school = ''
    if request.method == 'POST':
        school = request.POST.get('school')
        capacity = request.POST.get('capacity')

        item = Classes(school=school, capacity=capacity, user=request.user)
        item.save()


    class_id = dict()
    class_list = Classes.objects.filter(user=request.user)
    unique = list(dict.fromkeys([i.school for i in class_list]))

    for i in unique:
        filter_list = list(class_list.filter(school=i))
        for j in filter_list:
            index = (unique.index(j.school) + 1) * 100 + filter_list.index(j) + 1
            class_id[j.pk] = index

    classes = [(i,class_id[i.pk]) for i in class_list]
    classes.sort(key=lambda x: x[1])

    selected_schools = [
        i.school for i in Schools.objects.filter(user=request.user)]

    return render(request, 'class_selection.html', {'selected_schools': selected_schools,
                                                    'class_list': classes}
                                                    )


@login_required
def table(request):
    id = ''
    if request.method == 'POST':

        school = request.POST.get("school")
        course = request.POST.get("course")
        num_of_session = request.POST.get("num_of_session")
        teacher = request.POST.get("teacher")
        first_day = request.POST.get("first_day")
        if num_of_session == 'دو روز در هفته':
            second_day = request.POST.get("second_day")
            second_start_time = request.POST.get("second_start_time")
            second_end_time = request.POST.get("second_end_time")
        else:
            second_day = ""
            second_start_time = ""
            second_end_time = ""

        first_start_time = request.POST.get("first_start_time")
        first_end_time = request.POST.get("first_end_time")
        signup_capacity = request.POST.get("signup_capacity")
        id = request.POST.get("id")

        if id:
            Table.objects.filter(pk=id).update(school=school, course=course,
                                               num_of_session=num_of_session, teacher=teacher,
                                               first_day=first_day, second_day=second_day,
                                               first_start_time=first_start_time, first_end_time=first_end_time,
                                               second_start_time=second_start_time, second_end_time=second_end_time,
                                               signup_capacity=signup_capacity)

        else:
            item = Table(user=request.user, school=school, course=course,
                         num_of_session=num_of_session, teacher=teacher,
                         first_day=first_day, second_day=second_day,
                         first_start_time=first_start_time, first_end_time=first_end_time,
                         second_start_time=second_start_time, second_end_time=second_end_time,
                         signup_capacity=signup_capacity)
            item.save()

    table_list = Table.objects.filter(user=request.user)
    selected_schools = [
        i.school for i in Schools.objects.filter(user=request.user)]


    return render(request, 'table.html', {'selected_schools': selected_schools,
                                          'table_list': table_list,
                                          'time_list': time_list,
                                          'id': id})


@login_required
def create_excel(request):

    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + \
        request.user.username + '.xlsx'
    class_list = Classes.objects.filter(user=request.user)
    table_list = Table.objects.filter(user=request.user)

    unique = list(dict.fromkeys([i.school for i in class_list]))

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Data"
    ws3 = wb.create_sheet("Capacity")

    columns = [('a1', 'ردیف'), ('b1', 'دانشکده'), ('c1', 'نام درس'), ('d1', 'تعداد جلسات در هفته'), ('e1', 'نام استاد'), ('f1', 'اولین روز درس'),
               ('g1', 'ساعت شروع'), ('h1', 'ساعت پایان'), ('i1', 'دومین روز درس'), ('j1', 'ساعت شروع'), ('k1', 'ساعت پایان'), ('l1', 'ظرفیت ثبت نامی احتمالی درس')]

    for i in columns:
        ws1[i[0]].value = i[1]
        ws1[i[0]].fill = PatternFill("solid", fgColor='E1C39B')

    ws3['a1'].value = 'شماره کلاس'
    ws3['b1'].value = 'نام دانشکده'
    ws3['c1'].value = 'ظرفیت کلاس'

    row = 2
    for object in table_list:
        ws1['a' + str(row)].value = row - 1
        ws1['b' + str(row)].value = object.school
        ws1['c' + str(row)].value = object.course
        ws1['d' + str(row)].value = object.num_of_session
        ws1['e' + str(row)].value = object.teacher
        ws1['f' + str(row)].value = object.first_day
        if object.first_day:
            ws1['g' + str(row)].value = '"' + object.first_start_time + '"'
            ws1['h' + str(row)].value = '"' + object.first_end_time + '"'
        if object.num_of_session == 'دو روز در هفته':
            ws1['i' + str(row)].value = object.second_day
            ws1['j' + str(row)].value = '"' + object.second_start_time + '"'
            ws1['k' + str(row)].value = '"' + object.second_end_time + '"'
        ws1['l' + str(row)].value = int(object.signup_capacity)
        row = row + 1

    row = 2

    for i in unique:
        filter_list = list(class_list.filter(school=i))
        for j in filter_list:
            index = (unique.index(j.school) + 1) * \
                100 + filter_list.index(j) + 1
            ws3['a' + str(row)].value = index
            ws3['b' + str(row)].value = j.school
            ws3['c' + str(row)].value = int(j.capacity)
            row = row + 1

    wb.save(response)
    return response


@login_required
def delete_table(request, pk):
    item = get_object_or_404(Table, pk=pk)
    item.delete()

    return redirect('table')


@login_required
def delete_class(request,pk):
    item = get_object_or_404(Classes, pk=pk)
    item.delete()

    return redirect('class_selection')


@login_required
def classDeleteAll(request):
    Classes.objects.filter(user=request.user).all().delete()
    return redirect('class_selection')


@login_required
def updateClass(request, pk):
    item = get_object_or_404(Classes, pk=pk)
    item.capacity = request.POST.get('capacity')
    item.school = request.POST.get('school')

    item.save()
    return redirect('class_selection')


@login_required
def upload(request):

    form = UploadFileForm()
    upload_list = Table.objects.filter(user=request.user, upload=True)

    if request.method == 'POST':
        form = UploadFileForm(request.POST, request.FILES)

        if form.is_valid():

            file = request.FILES['file']
            wb = load_workbook(file)
            ws = wb.active
            row_list = list()
            max_row = ws.max_row

            for i in range(2, max_row + 1):
                if ws['a' + str(i)].value:
                    row_list.append(ws['a' + str(i)].row)

            item_list = list()

            for i in range(len(row_list)):
                if not i == len(row_list) - 1:
                    item = (row_list[i], row_list[i + 1] - row_list[i])
                    item_list.append(item)
                else:
                    item = (row_list[i], max_row - row_list[i] + 1)
                    item_list.append(item)

            ###################################################################
            checked_continue = request.POST.get('continue')
            if not checked_continue == 'True':
                item = Table.objects.filter(
                    user=request.user, upload=True).all().delete()

            for i in item_list:

                time_list = list()

                for j in range(i[1]):
                    if ws['O' + str(i[0] + j)].value:
                        time_list.append(ws['O' + str(i[0] + j)].value)

                ##########################################################

                if len(time_list) == 3:
                    first_day = time_list[0]
                    second_day = time_list[1]

                elif len(time_list) == 2:
                    if 'درس' in time_list[0]:
                        first_day = time_list[0]
                    else:
                        first_day = ''
                    if 'درس' in time_list[1]:
                        second_day = time_list[1]
                    else:
                        second_day = ''

                elif len(time_list) == 1:
                    if time_list[0] == None:
                        first_day = ''
                        second_day = ''
                    else:
                        if 'درس' in time_list[0]:
                            first_day = time_list[0]
                        else:
                            first_day = ''
                        second_day = ''

                ##############################################################
                def sp(x):
                    if not x == '':
                        if len(x.split(' ')) == 4:
                            if x.split(' ')[1] == 'سه':
                                return 'سه‌شنبه'
                            elif x.split(' ')[1] == 'یک' or x.split(' ')[1] == 'يك':
                                return sunday

                            else:
                                return (x.split(' ')[1] + x.split(' ')[2])
                        else:
                            return x.split(' ')[1]
                    else:
                        return x

                new_first_day = sp(first_day)
                new_second_day = sp(second_day)

                if not new_first_day == '':
                    first_start_time = first_day.split(' ')[-1].split('-')[0]
                    first_end_time = first_day.split(' ')[-1].split('-')[1]
                else:
                    first_start_time = ''
                    first_end_time = ''

                if not new_second_day == '':
                    second_start_time = second_day.split(' ')[-1].split('-')[0]
                    second_end_time = second_day.split(' ')[-1].split('-')[1]
                else:
                    second_start_time = ''
                    second_end_time = ''

                def time_format(x):
                    if x:
                        if len(x.split(":")[0]) == 2:
                            if x.split(":")[0][0] == '0':
                                return x[1:]
                            else:
                                return x
                        else:
                            return x
                    else:
                        return x

                first_start_time = time_format(first_start_time)
                first_end_time = time_format(first_end_time)
                second_start_time = time_format(second_start_time)
                second_end_time = time_format(second_end_time)

                ##########################################################

                if first_day and second_day:
                    num_of_session = 'دو روز در هفته'
                elif first_day:
                    num_of_session = 'یک روز در هفته'
                else:
                    num_of_session = ''
                item = Table(user=request.user,
                             school=ws['E' + str(i[0])].value,
                             course=ws['G' + str(i[0])].value,
                             teacher=ws['N' + str(i[0])].value,
                             signup_capacity=ws['J' + str(i[0])].value,
                             num_of_session=num_of_session,
                             first_day=new_first_day,
                             second_day=new_second_day,
                             first_start_time=first_start_time,
                             first_end_time=first_end_time,
                             second_start_time=second_start_time,
                             second_end_time=second_end_time,
                             upload=True,
                             )
                item.save()

                upload_list = Table.objects.filter(
                    user=request.user, upload=True)

    return render(request, 'upload.html', {'form': form,
                                           'upload_list': upload_list})

def guide(request):
    return render(request,'guide.html')

@login_required
def mark_table(request):
    mark_list = request.POST.getlist('checkbox')
    Table.objects.filter(pk__in=mark_list).delete()

    return redirect('upload')


@login_required
def delete_all(request):
    Table.objects.filter(user=request.user).all().delete()

    return redirect('table')


@login_required
def delete_uploaded(request):
    Table.objects.filter(user=request.user, upload=True).all().delete()

    return redirect('table')


@login_required
def delete_manuals(request):
    Table.objects.filter(user=request.user, upload=False).all().delete()

    return redirect('table')


@login_required
def finalize(request):
    message = str()
    if request.method == 'POST':

        def Float(x):
            return float(x)

        ww = list(map(Float,request.POST.getlist('ww')))
        if sum(ww[i] for i in range(len(ww))) == 0:
            message = 'هر سه ضريب نمي‌تواند صفر باشد.'
        else:

            courses = Table.objects.filter(user=request.user)
            classes = Classes.objects.filter(user=request.user)
            if len(classes) == 0:
                message = 'هيچ كلاسي وارد نشده است.'
            else:

                I = range(len(courses))
                K = range(len(classes))
                J = range(322)


            ########################## HH ########################

                HH = [[0 for i in J] for y in courses]


                def day_compiler(x):
                    if x == 'شنبه':
                        return 0
                    elif x == 'یکشنبه' or x == 'يكشنبه' or x == 'يك‌شنبه' or x == 'یک‌شنبه':
                        return 1
                    elif x == 'دوشنبه':
                        return 2
                    elif x == 'سه‌شنبه':
                        return 3
                    elif x == 'چهارشنبه':
                        return 4
                    elif x == 'پنجشنبه':
                        return 5
                    elif x == 'جمعه':
                        return 6

                def time_index(x):
                    return time_list.index(x)

                for i in I:
                    try:
                        start = day_compiler(courses[i].first_day)*46+time_index(courses[i].first_start_time)
                        end = day_compiler(courses[i].first_day)*46+time_index(courses[i].first_end_time)
                        HH[i][start:end] = [1] * (end - start)

                        if courses[i].second_day:
                            start = day_compiler(courses[i].second_day)*46+time_index(courses[i].second_start_time)
                            end = day_compiler(courses[i].second_day)*46+time_index(courses[i].second_end_time)
                            HH[i][start:end] = [1] * (end - start)
                    except:
                        message = 'زمان شروع يا پايان درس ' + courses[i].course + ' با شماره ' + str(i+1) +'در بازه زماني تعيين شده قرار ندارد جهت اصلاح آن به صورت دستي از لينك زير استفاده كنيد.'
                        break

                if not message:


                #################### HHH ######################

                    HHH = list()
                    for i in I:
                        HHH.append(sum(HH[i][j] for j in J))


                ###################### classcourse ######################

                    classcourse = [[0 for k in K] for i in courses]

                    for (i,k) in product(I,K):
                        if courses[i].school == classes[k].school:
                            classcourse[i][k] = 1

                ########################## capa ##########################

                    capa = [i.capacity for i in classes]

                ########################  reg  ##########################

                    reg = [i.signup_capacity for i in courses]

                ######################### ww ##############################

                    # ww = [0,0,0]

                ####################### variables ##########################

                    m = Model()

                    x = [ [m.add_var(name="x({},{})".format(courses[i].course,k+1),var_type=BINARY) for k in K] for i in I ]
                    r = [ m.add_var(name="r({})".format(k+1), var_type=BINARY) for k in K]

                    y = [ [m.add_var(name="y({},{})".format(courses[i].course,k+1),lb=0) for k in K] for i in I]
                    rr = m.add_var(name="rr",lb=0)
                    cls = m.add_var(name="cls",lb=0)
                    brk = m.add_var(name="brk",lb=0)


                ########################### objective function ###############################

                    m.objective = minimize( ww[0]*rr + ww[1]*brk - ww[2]*cls )

                ########################### Constraints ###############################

                    for (j,k) in product(J,K):
                        m += xsum( HH[i][j]*x[i][k] for i in I ) <= 1

                    for (i,k) in product(I,K):
                        m += y[i][k] >= (int(reg[i])-int(capa[k])) * x[i][k]

                    for i in I:
                        if HHH[i] > 0:
                            m += xsum( x[i][k] for k in K ) == 1

                    for (i,k) in product(I,K):
                        m += x[i][k] <= r[k]

                    m += rr == xsum( r[k] for k in K )

                    m += cls == xsum( classcourse[i][k]*x[i][k] for i in I for k in K )

                    m+= brk == xsum( y[i][k] for i in I for k in K )

                    ###########################################################
                    status = m.optimize()
                    text = str()
                    for v in m.vars:
                        text += '<li>'+ v.name + ' = ' + str(v.x) +'</li>'
                    text += '<li>'+'objective function' + ' = '  + str(m.objective_value) +'</li>'
                    if status == OptimizationStatus.INFEASIBLE:
                        message = 'جواب بهينه وجود ندارد.'

                    else:
                        response = HttpResponse(content_type='application/vnd.ms-excel')
                        response['Content-Disposition'] = 'attachment; filename=' + \
                            'results' + '.xlsx'

                    ####################### Tiemtable #########################

                        solution_list = list()
                        unique = list(dict.fromkeys([i.school for i in classes]))

                        for i in range(len(HH)):
                            for k in range(len(HH[i])):
                                if HH[i][k] == 1:
                                    try:
                                        HH[i][k] = unique.index(courses[i].school) + 1
                                    except:
                                        pass
                        # print(HH[1])
                        class_id = dict()
                        index_list = list()

                        for i in unique:
                            filter_list = list(classes.filter(school=i))
                            for j in filter_list:
                                index = (unique.index(j.school) + 1) * 100 + filter_list.index(j) + 1
                                index_list.append(index)
                                class_id[j.pk] = index

                        for (i,k) in product(I,K):
                            if x[i][k].x > 0:
                                solution_list.append((courses[i].course,class_id[classes[k].pk],i))

                        wb = Workbook()
                        ws1 = wb.active
                        ws1.title = 'Timetable'

                        days = [('شنبه','E1C39B','ECE98A'),
                                ('يكشنبه','A1DD8B','A2F069'),
                                ('دوشنبه','8BDDD0','69EAF0'),
                                ('سه‌شنبه','8BC9DD','69C1F0'),
                                ('چهارشنبه','8AA3EC','6992F0'),
                                ('پنجشنبه','E1C39B','ECE98A'),
                                ('جمعه','A1DD8B','A2F069')]
                        def template(x):
                            col = 3
                            for day in days:
                                for i in range(len(time_list)-1):
                                    x.cell(row=1, column=col, value=day[0]).fill = PatternFill("solid", fgColor=day[1])
                                    x.cell(row=2, column=col, value=time_list[i]+'-'+time_list[i+1]).fill = PatternFill("solid", fgColor=day[2])
                                    col += 1

                        template(ws1)

                        ws1.cell(row=2, column=1, value='رديف').fill = PatternFill("solid", fgColor='F0BD69')
                        ws1.cell(row=2, column=2, value='نام درس').fill = PatternFill("solid", fgColor='F0BD69')

                        for i in range(len(HH)):
                            col = 3
                            ws1.cell(row=i+3, column=1, value=i+1)
                            ws1.cell(row=i+3, column=2, value=solution_list[i][0])
                            for j in range(len(HH[i])):
                                if HH[i][j] != 0:
                                    ws1.cell(row=i+3, column=col, value=solution_list[i][1])
                                else:
                                    ws1.cell(row=i+3, column=col, value=0)
                                col += 1


                        #################### Classes #####################
                        ws2 = wb.create_sheet('Classes')
                        template(ws2)

                        for i in range(len(index_list)):
                            col = 3
                            ws2.cell(row=i+3, column=1, value=index_list[i])
                            items = [HH[j[2]] for j in solution_list if j[1] == index_list[i]]
                            column_sums = [sum([row[i] for row in items]) for i in J]
                            for j in range(len(column_sums)):
                                ws2.cell(row=i+3, column=col, value=column_sums[j])
                                col += 1

                        ################### variables ####################


                        ################### X(i,k) ####################

                        ws3 = wb.create_sheet('x(i,k)')

                        ws3.cell(row=1, column=1, value='x(i,k)')

                        for k in K:
                            ws3.cell(row=1,column=k+2,value=class_id[classes[k].pk])
                        for i in I:
                            ws3.cell(row=i+2,column=1,value=courses[i].course)

                        for (i,k) in product(I,K):
                            ws3.cell(row=i+2,column=k+2,value=x[i][k].x)

                        ################### r(k) ####################

                        ws4 = wb.create_sheet('r(k)')
                        ws4.cell(row=1, column=1, value='r(k)')

                        r_solution = list()

                        for k in K:
                            r_solution.append((r[k].x,class_id[classes[k].pk]))
                        r_solution.sort(key=lambda x:x[1])

                        k = 0
                        for i in r_solution:
                            ws4.cell(row=k+2, column=1, value=i[1])
                            ws4.cell(row=k+2, column=2, value=i[0])
                            k += 1

                        # ################### Y(i,k) ####################

                        ws5 = wb.create_sheet('y(i,k)')

                        ws5.cell(row=1, column=1, value='y(i,k)')


                        for k in K:
                            ws5.cell(row=1,column=k+2,value=class_id[classes[k].pk])
                        for i in I:
                            ws5.cell(row=i+2,column=1,value=courses[i].course)

                        for (i,k) in product(I,K):
                            ws5.cell(row=i+2,column=k+2,value=y[i][k].x)
                        #
                        # ################### rr ####################

                        ws6 = wb.create_sheet('rr,cls,brk,OBJ')


                        ws6.cell(row=1, column=1, value='rr')
                        ws6.cell(row=2, column=1, value=rr.x)

                        # ################### cls ####################
                        #
                        ws6.cell(row=1, column=3, value='cls')
                        ws6.cell(row=2, column=3, value=cls.x)
                        #
                        # ################### brk ####################
                        #
                        ws6.cell(row=1, column=5, value='brk')
                        ws6.cell(row=2, column=5, value=brk.x)
                        #
                        # ############# objective fucntion ###############
                        #
                        ws6.cell(row=1, column=7, value='objective function')
                        ws6.cell(row=2, column=7, value=m.objective_value)

                        def center(x):
                            for col in x.columns:
                                for cell in col:
                                    cell.alignment = styles.Alignment(horizontal='center')

                        center(ws1)
                        center(ws2)
                        center(ws3)
                        center(ws4)
                        center(ws5)
                        center(ws6)


                        wb.save(response)
                        return response



    schools = Schools.objects.filter(user=request.user)
    classes = Classes.objects.filter(user=request.user)
    courses = Table.objects.filter(user=request.user)

    return render(request, 'finalize.html', {'schools': schools,
                                             'classes': classes,
                                             'courses': courses,
                                             'message': message
                                             })
